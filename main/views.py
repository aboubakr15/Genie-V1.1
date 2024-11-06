from io import BytesIO
from openpyxl.utils import get_column_letter
from django.shortcuts import render, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect
from django.core.exceptions import ValidationError
from leads.forms import UploadSheetsForm
from .utils import has_valid_contact, is_valid_phone_number, clean_company_name, filter_companies, get_string_value, get_lead_related_data
from .models import (Lead, Sheet, LeadContactNames, LeadEmails, LeadPhoneNumbers, Log, LeadsAverage, UserLeader, FilterWords,
                    FilterType, LeadsColors, SalesLog, LeadTerminationCode, Notification, TaskLog, TerminationCode)
from .forms import AutoFillForm, FilterWordsForm # ImportSheetsForm  # For mass importing 
import os, logging, pandas as pd
from django.core.exceptions import ObjectDoesNotExist
from datetime import datetime, timezone as dt_timezone, timedelta
from operations_team_leader.forms import LeadForm
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden
from openpyxl.styles import PatternFill
from django.contrib.auth.decorators import user_passes_test, login_required
from main.custom_decorators import is_in_group
from django.http import JsonResponse
from django.core.exceptions import PermissionDenied
from django.db.models import Sum, Q
from django.contrib.auth.models import User, Group
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
import json, pytz
import random, string

# For Mass Import
# from concurrent.futures import ThreadPoolExecutor
# from django.db import transaction

def generate_random_string(length=5):
    """Generate a random string of fixed length."""
    letters = string.ascii_letters  # Use both uppercase and lowercase letters
    return ''.join(random.choice(letters) for i in range(length))

logger = logging.getLogger('custom')

def index(request):
    if request.user.is_authenticated: 
        return redirect(f"/{request.user.groups.first().name}")
    return redirect("main:login")


def login_view(request):
    context = {"error": ""}
    
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()

        try:
            if not username or not password:
                raise ValidationError("Username and password are required.")
            
            if len(username) < 3 or len(username) > 20:
                raise ValidationError("Username must be between 3 and 20 characters long.")
            
            if len(password) < 8:
                raise ValidationError("Password must be at least 8 characters long.")

        except ValidationError as e:
            context['error'] = str(e)
            return render(request, "main/login.html", context=context)

        user = authenticate(request, username=username, password=password)

        if user is not None:
            send_cb_date_notifications()

            group = user.groups.first()
            if group:
                login(request, user)
                Log.objects.create(message=f"User [{user.username}] Logged in successfully")
                return redirect(f"/{group.name}")
            else:
                context['error'] = 'Oops, User has no group. Please ask the admin to add you to one.'
        else:
            logger.warning(f"Failed login attempt for {username}")
            context['error'] = 'Wrong username or password.'

    return render(request, "main/login.html", context=context)


def logout_view(request):
    username = request.user.username  # Get username before logging out
    logout(request)
    Log.objects.create(message=f"User [{username}] has logged out of the system")
    return redirect("/login/")


def sheet_list(request):
    term = request.GET.get('q', '')
    sheets = Sheet.objects.filter(name__icontains=term)
    results = [{'id': sheet.id, 'text': sheet.name} for sheet in sheets]
    return JsonResponse({'items': results})


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager") or is_in_group(user, "Sales_manager"))
def lead_details(request, pk):
    # Get the lead instance based on the primary key (pk)
    lead = get_object_or_404(Lead, pk=pk)
    
    # Retrieve all related phone numbers, emails, and contact names
    phone_numbers = lead.leadphonenumbers_set.all()
    emails = lead.leademails_set.all()
    contact_names = lead.leadcontactnames_set.all()

    # Use a set to remove duplicates based on value
    unique_phone_numbers = {pn.value: pn for pn in phone_numbers}.values()
    unique_emails = {em.value: em for em in emails}.values()
    unique_contact_names = {cn.value: cn for cn in contact_names}.values()
    
    # Retrieve sheets associated with the lead
    sheets = lead.sheets.all()
    
    # Get the user's group name
    group_name = None
    if request.user.groups.exists():
        group_name = request.user.groups.first().name

    # Determine the template path based on group
    if group_name == 'operations_manager':
        template_path = 'operations_manager/lead_details.html'
    elif group_name == 'operations_team_leader':
        template_path = 'operations_team_leader/lead_details.html'
    elif group_name == 'sales_manager':
        template_path = 'sales_manager/lead_details.html'
    else:
        return HttpResponseForbidden("You do not have permission to view this page.")

    # Pass all the details to the template
    data = {
        'lead': lead,
        'phone_numbers': unique_phone_numbers,
        'emails': unique_emails,
        'contact_names': unique_contact_names,
        'sheets': sheets,
    }
    
    return render(request, template_path, data)


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def add_lead(request):
    group_name = None
    if request.user.groups.exists():
        group_name = request.user.groups.first().name

    if request.method == 'POST':
        form = LeadForm(request.POST)
        if form.is_valid():
            lead = form.save(commit=False)

            # Debugging
            print("Selected Sheets:", form.cleaned_data.get('sheets'))

            sheets = form.cleaned_data.get('sheets')
            if not sheets:
                messages.error(request, "Please select at least one sheet.")
                return render(request, f'{group_name}/add_lead.html', {
                    'form': form,
                })

            lead.save()

            phone_numbers = form.cleaned_data.get('phone_numbers')
            emails = form.cleaned_data.get('emails')
            contact_names = form.cleaned_data.get('contact_names')

            for phone_number in phone_numbers.split(','):
                phone_number = phone_number.strip()
                if phone_number:
                    LeadPhoneNumbers.objects.get_or_create(lead=lead, value=phone_number, sheet=sheets.first())

            for email in emails.split(','):
                email = email.strip()
                if email:
                    LeadEmails.objects.get_or_create(lead=lead, value=email, sheet=sheets.first())

            for contact_name in contact_names.split(','):
                contact_name = contact_name.strip()
                if contact_name:
                    LeadContactNames.objects.get_or_create(lead=lead, value=contact_name, sheet=sheets.first())

            # Associate lead with all selected sheets
            for sheet in sheets:
                sheet.leads.add(lead)

            messages.success(request, "Lead added successfully!")

            if group_name == 'operations_manager':
                return redirect('operations_manager:index')
            elif group_name == 'operations_team_leader':
                return redirect('operations_team_leader:index')
            else:
                return HttpResponseForbidden("You do not have permission to perform this action.")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = LeadForm()

    return render(request, f"{group_name}/add_lead.html", {
        'form': form,
    })


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def edit_lead(request, pk):
    group_name = None
    if request.user.groups.exists():
        group_name = request.user.groups.first().name
    logger.debug(f"User group: {group_name}")

    lead = get_object_or_404(Lead, id=pk)

    if request.method == 'POST':
        form = LeadForm(request.POST, instance=lead)
        if form.is_valid():
            lead = form.save(commit=False)

            sheets = form.cleaned_data.get('sheets')
            if not sheets:
                messages.error(request, "Please select at least one sheet.")
                return render(request, f'{group_name}/edit_lead.html', {
                    'form': form,
                })

            lead.save()

            # Update Phone Numbers
            current_phone_numbers = LeadPhoneNumbers.objects.filter(lead=lead).values_list('value', flat=True)
            new_phone_numbers = {phone_number.strip() for phone_number in form.cleaned_data.get('phone_numbers').split(',') if phone_number.strip()}

            # Delete removed phone numbers
            LeadPhoneNumbers.objects.filter(lead=lead).exclude(value__in=new_phone_numbers).delete()
            # Add new phone numbers
            for phone_number in new_phone_numbers.difference(current_phone_numbers):
                for sheet in sheets:
                    LeadPhoneNumbers.objects.get_or_create(lead=lead, value=phone_number, sheet=sheet)

            # Update Emails
            current_emails = LeadEmails.objects.filter(lead=lead).values_list('value', flat=True)
            new_emails = {email.strip() for email in form.cleaned_data.get('emails').split(',') if email.strip()}

            # Delete removed emails
            LeadEmails.objects.filter(lead=lead).exclude(value__in=new_emails).delete()
            # Add new emails
            for email in new_emails.difference(current_emails):
                for sheet in sheets:
                    LeadEmails.objects.get_or_create(lead=lead, value=email, sheet=sheet)

            # Update Contact Names
            current_contact_names = LeadContactNames.objects.filter(lead=lead).values_list('value', flat=True)
            new_contact_names = {contact_name.strip() for contact_name in form.cleaned_data.get('contact_names').split(',') if contact_name.strip()}

            # Delete removed contact names
            LeadContactNames.objects.filter(lead=lead).exclude(value__in=new_contact_names).delete()
            # Add new contact names
            for contact_name in new_contact_names.difference(current_contact_names):
                for sheet in sheets:
                    LeadContactNames.objects.get_or_create(lead=lead, value=contact_name, sheet=sheet)

            # Update Sheets
            current_sheets = lead.sheets.all()
            # Remove the lead from sheets that were deselected
            for sheet in current_sheets.difference(sheets):
                sheet.leads.remove(lead)
            # Add the lead to newly selected sheets
            for sheet in sheets.difference(current_sheets):
                sheet.leads.add(lead)

            messages.success(request, "Lead updated successfully!")

            if group_name == 'operations_manager':
                return redirect('operations_manager:index')
            elif group_name == 'operations_team_leader':
                return redirect('operations_team_leader:index')
            else:
                return HttpResponseForbidden("You do not have permission to perform this action.")
        else:
            # Log form and formset errors
            logger.debug("Form or formsets are invalid")
            messages.error(request, "Please correct the errors below.")
            for error in form.errors:
                messages.error(request, f"Form Error: {error}")
                logger.debug(f"Form Error: {error}")

    else:
        # Initialize form with existing lead data
        form = LeadForm(instance=lead)
        # Set initial values for other fields
        form.fields['sheets'].initial = lead.sheets.all()  # Make sure to use `all()` to get a queryset of sheets
        form.fields['phone_numbers'].initial = ', '.join(LeadPhoneNumbers.objects.filter(lead=lead).values_list('value', flat=True))
        form.fields['emails'].initial = ', '.join(LeadEmails.objects.filter(lead=lead).values_list('value', flat=True))
        form.fields['contact_names'].initial = ', '.join(LeadContactNames.objects.filter(lead=lead).values_list('value', flat=True))

    return render(request, f"{group_name}/edit_lead.html", {
        'form': form,
    })


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def delete_lead(request, pk):
    # Fetch the lead instance or return a 404 error if not found
    lead = get_object_or_404(Lead, pk=pk)
    
    # Get the user's group name
    group_name = None
    if request.user.groups.exists():
        group_name = request.user.groups.first().name

    if request.method == 'POST':
        lead.sheets.clear()
        lead.delete()
        
        # Redirect based on user group
        if group_name == 'operations_manager':
            return redirect('operations_manager:index')
        elif group_name == 'operations_team_leader':
            return redirect('operations_team_leader:index')
        else:
            return HttpResponseForbidden("You do not have permission to perform this action.")

    # Determine the template path based on group
    if group_name == 'operations_manager':
        template_path = 'operations_manager/delete_lead.html'
    elif group_name == 'operations_team_leader':
        template_path = 'operations_team_leader/delete_lead.html'
    else:
        return HttpResponseForbidden("You do not have permission to view this page.")

    # Render the template with lead information
    return render(request, template_path, {'lead': lead})


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def auto_fill(request):

    template_name = 'operations_team_leader/auto_fill.html'  # Default template

    if request.user.groups.filter(name='operations_manager').exists():
        template_name = 'operations_manager/auto_fill.html'

    if request.method == 'POST':
        form = AutoFillForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            selected_sheet = form.cleaned_data.get('latest_sheet')

            try:
                # Load data
                if file.name.endswith('.xlsx'):
                    data = pd.read_excel(file, engine='openpyxl')
                elif file.name.endswith('.xls'):
                    data = pd.read_excel(file)
                elif file.name.endswith('.csv'):
                    data = pd.read_csv(file)
                else:
                    messages.error(request, "Unsupported file format.")
                    return render(request, template_name, {'form': form})

                # Ensure required columns are present
                required_columns = ['Company Name', 'Time Zone', 'Phone Number', 'Email', 'DM Name', 'Color']
                for col in required_columns:
                    if col not in data.columns:
                        data[col] = ''

                num_leads_total = len(data['Company Name'].unique())
                data['Company Name'] = data['Company Name'].map(clean_company_name)
                data = data[data['Company Name'].apply(filter_companies)]
                company_names = data['Company Name'].unique()

                leads = Lead.objects.filter(name__in=company_names)
                leads_dict = {lead.name: lead for lead in leads}
                num_leads_autofilled = len(leads)

                latest_leads_dict = {}

                # Only proceed with the latest sheet if it is selected
                if selected_sheet:
                    latest_leads = selected_sheet.leads.all()
                    latest_leads_dict = {lead.name: lead for lead in latest_leads}

                    # Add new rows to the data from the latest sheet
                    for lead in latest_leads:
                        phone_number, email, contact_name = get_lead_related_data(lead)
                        data = data[data['Company Name'] != lead.name]

                        new_row = pd.DataFrame([{
                            'Company Name': lead.name,
                            'Time Zone': lead.time_zone,
                            'Phone Number': phone_number,
                            'Email': email,
                            'DM Name': contact_name,
                            'Color': 'green'
                        }])

                        # Ensure the new_row has the same columns as data
                        new_row = new_row.reindex(columns=data.columns)

                        # Append new_row to the existing DataFrame
                        data = pd.concat([data, new_row], ignore_index=True)

                def fill_data(row):
                    company_name = row.get('Company Name')
                    if company_name not in latest_leads_dict:
                        lead = leads_dict.get(company_name)
                        if lead:
                            phone_number, email, contact_name = get_lead_related_data(lead)
                            row['Time Zone'] = lead.time_zone
                            row['Phone Number'] = phone_number
                            row['Email'] = email
                            row['DM Name'] = contact_name
                            row['Color'] = 'blue'
                        else:
                            row['Time Zone'] = ''
                            row['Phone Number'] = ''
                            row['Email'] = ''
                            row['DM Name'] = ''
                            row['Color'] = 'red'
                    return row

                data = data.apply(fill_data, axis=1)
                data.reset_index(drop=True, inplace=True)

                # Save file to server
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    data.to_excel(writer, index=False, sheet_name='Leads')
                    workbook = writer.book
                    worksheet = writer.sheets['Leads']

                    column_widths = {
                        'Company Name': 20,
                        'Time Zone': 15,
                        'Phone Number': 20,
                        'Email': 25,
                        'DM Name': 20,
                        'Color': 10
                    }

                    for col, width in column_widths.items():
                        col_idx = data.columns.get_loc(col) + 1
                        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

                    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

                    for idx, row in data.iterrows():
                        color = row['Color']
                        fill = green_fill if color == 'green' else blue_fill if color == 'blue' else red_fill
                        for col_num in range(1, len(row) + 1):
                            worksheet.cell(row=idx + 2, column=col_num).fill = fill

                output.seek(0)
                
                # Provide file download to the user
                filename = f"result_{file.name}"
                response = HttpResponse(
                    output,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}"'

                logger.info(f"[{request.user.username}] Made an AutoFill with [{num_leads_total}] total leads, [{num_leads_autofilled}] autofilled from the database.")
                Log.objects.create(message=f"[{request.user.username}] Made an AutoFill with [{num_leads_total}] total leads, [{num_leads_autofilled}] autofilled from the database.")

                return response

            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
                logger.error(f"Error processing file: {str(e)}", exc_info=True)
                return render(request, template_name, {'form': form})

    else:
        form = AutoFillForm()
        messages.error(request, "Form is not valid. Please check the form fields.")
        print(form.errors)  # Print form errors to terminal

    return render(request, template_name, {'form': form})


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def upload_sheet(request):
    template_name = "operations_team_leader/upload_sheet.html" if is_in_group(request.user, "operations_team_leader") else "operations_manager/upload_sheet.html"

    form = UploadSheetsForm()

    if request.method == 'POST':
        form = UploadSheetsForm(request.POST, request.FILES)
        if 'file' in request.FILES:
            file = request.FILES['file']

            if file.name.endswith('.xlsx'):
                data = pd.read_excel(file, engine='openpyxl', header=None)
            elif file.name.endswith('.xls'):
                data = pd.read_excel(file, header=None)
            elif file.name.endswith('.csv'):
                data = pd.read_csv(file, header=None)
            else:
                return HttpResponse("Unsupported file format.", status=400)

            # Check if the DataFrame is empty
            if data.empty:
                messages.error(request, "The uploaded file is empty.")
                return render(request, template_name, {"form": form})

            offset = timedelta(hours=3)  # Egypt is UTC+3
            egypt_timezone = pytz.timezone('Africa/Cairo')
            current_time_with_offset = datetime.now(egypt_timezone)

            # Ensure all headers are strings to avoid AttributeError
            
            def ensure_str(value):
                return str(value) if pd.notna(value) else ''

            # Determine if header row is present
            expected_headers = ['Company Name', 'Phone Number', 'Time Zone', 'Email', 'DM Name']
            extended_expected_headers = ['Company Name', 'Phone Number', 'Time Zone', 'Direct / Cell Number', 'Email', 'DM Name']

            first_row_as_str = data.iloc[0].apply(ensure_str) if not data.empty else pd.Series()

            # Check for headers using two possible expected formats
            has_header = all(header.lower() in [col.lower() for col in first_row_as_str if col] for header in expected_headers)
            has_extended_header = all(header.lower() in [col.lower() for col in first_row_as_str if col] for header in extended_expected_headers)

            if has_header or has_extended_header:
                # Use the first row as headers
                data.columns = data.iloc[0].apply(ensure_str).tolist()
                data = data[1:]  # Skip the header row

                # Handle missing headers by assigning names based on expected order
                if has_extended_header:
                    # Fill in missing extended headers
                    for i, header in enumerate(extended_expected_headers):
                        if i >= len(data.columns) or data.columns[i] == '' or data.columns[i].lower() not in [h.lower() for h in extended_expected_headers]:
                            data.columns[i] = header
                else:
                    # Fill in missing headers for default set
                    for i, header in enumerate(expected_headers):
                        if i >= len(data.columns) or data.columns[i] == '' or data.columns[i].lower() not in [h.lower() for h in expected_headers]:
                            data.columns[i] = header
            else:
                # Assign default column names based on order
                num_cols = data.shape[1]
                default_columns = ['Company Name', 'Phone Number', 'Time Zone', 'Direct / Cell Number', 'Email', 'DM Name']
                # Trim or extend default columns to match the actual number of columns in the DataFrame
                data.columns = default_columns[:num_cols] + [f"Column_{i}" for i in range(num_cols - len(default_columns))]

            # Clean & Filter company names
            if 'Company Name' in data.columns:
                data['Company Name'] = data['Company Name'].map(clean_company_name)
                data = data[data['Company Name'].apply(filter_companies)]


            random_suffix = generate_random_string(5)  # Generate a random suffix
            unique_sheet_name = f"{file.name}_{random_suffix}"  # Create a unique name

            # Create or get the sheet
            sheet, created = Sheet.objects.get_or_create(
                name=unique_sheet_name,
                defaults={'user': request.user, 'created_at': current_time_with_offset}
            )

            # Count new leads added
            new_leads_count = 0

            # Read leads and save to the database
            for _, row in data.iterrows():
                if not has_valid_contact(row):
                    continue

                company_name = row.get('Company Name', '')

                # Check if the lead already exists
                lead, created = Lead.objects.get_or_create(name=company_name)

                # If a new lead was created, increase the counter
                if created:
                    new_leads_count += 1

                # Add the time zone
                time_zone = get_string_value(row, 'Time Zone')
                if time_zone:
                    lead.time_zone = time_zone
                    lead.save()

                # Add the phone numbers if they're new
                phone_number = get_string_value(row, 'Phone Number')
                direct_cell_number = get_string_value(row, 'Direct / Cell Number') if 'Direct / Cell Number' in data.columns else None
                phone_numbers = ','.join(filter(None, [phone_number, direct_cell_number]))

                if phone_numbers:
                    for phone_number in phone_numbers.split(','):
                        phone_number = phone_number.strip()
                        if phone_number and is_valid_phone_number(phone_number):
                            try:
                                LeadPhoneNumbers.objects.get(lead=lead, sheet=sheet, value=phone_number)
                            except ObjectDoesNotExist:
                                LeadPhoneNumbers.objects.create(lead=lead, sheet=sheet, value=phone_number)
                        ## Add this part if u want to add org, gov, etc.. to the filter words automatically
                        # else:
                        #     # If the phone number is invalid & not ['nf', 'local'] add the company name to FilterWords
                        #     if company_name.lower() not in ['nf', 'local']:
                        #         phone_filter_type = FilterType.objects.get(name='phone_number')
                        #         filter_word, created = FilterWords.objects.get_or_create(word=company_name)

                        #         # Assign the existing filter type to the FilterWords instance
                        #         filter_word.filter_types.add(phone_filter_type)

                # Add the email if it's new
                email = get_string_value(row, 'Email')
                if email:
                    try:
                        LeadEmails.objects.get(lead=lead, sheet=sheet, value=email)
                    except ObjectDoesNotExist:
                        LeadEmails.objects.create(lead=lead, sheet=sheet, value=email)

                # Add the contact name if it's new
                contact_name = get_string_value(row, 'DM Name')
                if contact_name:
                    try:
                        LeadContactNames.objects.get(lead=lead, sheet=sheet, value=contact_name)
                    except ObjectDoesNotExist:
                        LeadContactNames.objects.create(lead=lead, sheet=sheet, value=contact_name)

                # Handle color if the 'Color' column exists
                if 'Color' in data.columns:
                    color = get_string_value(row, 'Color')
                    if color and color.lower() in ['white', 'green', 'blue', 'red']:
                        LeadsColors.objects.create(lead=lead, sheet=sheet, color=color.lower())


                # Link the lead to the sheet
                sheet.leads.add(lead)

            # Mark sheet as approved
            sheet.is_approved = True
            sheet.save()

            # If there are new leads, save the count in LeadsAverage
            if new_leads_count > 0:
                LeadsAverage.objects.create(
                    user=request.user,
                    sheet=sheet,
                    count=new_leads_count,
                    created_at=current_time_with_offset
                )

            # Add a success message
            messages.success(request, "Sheet imported successfully.")

            logger.info(f"{request.user.username} uploaded a sheet into the database containing {new_leads_count} new leads.")
            Log.objects.create(message=f"{request.user.username} uploaded a sheet into the database containing {new_leads_count} new leads.")

            return render(request, template_name, {"form": form})
        else:
            messages.error(request, "Please upload a file.")
            return render(request, template_name, {"form": form})
    else:
        logger.error("The specified file does not exist.")
        return render(request, template_name, {"form": form})


@user_passes_test(lambda user: is_in_group(user, "leads") or is_in_group(user, "operations_team_leader"))
def leads_average_view(request):
    user = request.user
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Get the date range
    start_date_obj, end_date_obj = None, None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        except ValueError:
            start_date_obj = None

    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        except ValueError:
            end_date_obj = None

    # Case for "leads" group
    if is_in_group(user, "leads"):
        leads_averages = LeadsAverage.objects.filter(user=user)
        if start_date_obj:
            leads_averages = leads_averages.filter(created_at__gte=start_date_obj)
        if end_date_obj:
            leads_averages = leads_averages.filter(created_at__lte=end_date_obj)

        # Calculate total average for the user
        total_leads_average = leads_averages.aggregate(Sum('count'))['count__sum'] or 0
        template_path = 'leads/leads_average.html'

        return render(request, template_path, {
            'leads_average': total_leads_average,
            'leads_averages': leads_averages,
            'start_date': start_date or '',
            'end_date': end_date or '',
        })

    # Case for "operations_team_leader" group
    elif is_in_group(user, "operations_team_leader"):
        # Get the team members of the leader (the user)
        team_members = UserLeader.objects.filter(leader=user).values_list('user', flat=True)
        team_leads_averages = LeadsAverage.objects.filter(user__in=team_members)
        leader_averages = LeadsAverage.objects.filter(user=user)

        if start_date_obj:
            team_leads_averages = team_leads_averages.filter(created_at__gte=start_date_obj)
            leader_averages = leader_averages.filter(created_at__gte=start_date_obj)

        if end_date_obj:
            team_leads_averages = team_leads_averages.filter(created_at__lte=end_date_obj)
            leader_averages = leader_averages.filter(created_at__lte=end_date_obj)

        # Calculate team averages
        team_averages = team_leads_averages.values('user').annotate(total_average=Sum('count'))
        user_averages = {}
        for avg in team_averages:
            user_id = avg['user']
            user = User.objects.get(id=user_id)
            user_averages[user.username] = avg['total_average']

        leader_total_average = leader_averages.aggregate(Sum('count'))['count__sum'] or 0
        total_team_average = leader_total_average + sum(user_averages.values())

        # Prepare data for rendering
        context = {
            'user_averages': {
                'team_members': user_averages,
                'leader': leader_total_average
            },
            'total_team_average': total_team_average,
            'leads_averages': team_leads_averages.union(leader_averages),
            'start_date': start_date or '',
            'end_date': end_date or '',
        }

        template_path = 'operations_team_leader/leads_average.html'
        return render(request, template_path, context)

    # Not authorized
    else:
        raise PermissionDenied("You are not authorized to view this page.")


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def manage_filter_words(request):

    template_name = "operations_team_leader/manage_filter_words.html" if is_in_group(request.user, "operations_team_leader") else "operations_manager/manage_filter_words.html"
    redirect_name = "operations_team_leader:manage-filter-words" if is_in_group(request.user, "operations_team_leader") else "operations_manager:manage-filter-words"

    query = request.GET.get('q', '')  # Get search query from the request
    filter_words_list = FilterWords.objects.all()

    if query:
        filter_words_list = filter_words_list.filter(word__icontains=query)  # Filter words by the search query

    filter_words_list = filter_words_list.order_by("-id")
    paginator = Paginator(filter_words_list, 30)
    page_number = request.GET.get('page')
    filter_words = paginator.get_page(page_number)

    if request.method == 'POST':
        form = FilterWordsForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(redirect_name)
    else:
        form = FilterWordsForm()

    return render(request, template_name, {
        'filter_words': filter_words,
        'form': form,
        'query': query,  # Pass the query to the template to keep the search term
    })


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def delete_filter_word(request, word_id):
    redirect_name = "operations_team_leader:manage-filter-words" if is_in_group(request.user, "operations_team_leader") else "operations_manager:manage-filter-words"
    filter_word = get_object_or_404(FilterWords, id=word_id)
    filter_word.delete()  # Delete the word
    return redirect(redirect_name)  # Redirect to the same page



@csrf_exempt
def log_inactivity(request):
    if request.method == 'POST':
        if lambda user: is_in_group(user, "sales_team_leader") or is_in_group(user, "sales"):
            data = json.loads(request.body)
            message = data.get('message', 'User inactive')
            user = request.user


        if user.groups.filter(name__in=['sales', 'sales_team_leader']).exists():
            # Log the message into SalesLog
            SalesLog.objects.create(
                message=message,
                date=timezone.now(),
                user=user
            )
            Log.objects.create(message=f"{request.user.username} was Inactive for 5 minutes")


            return JsonResponse({"status": "success", "message": "Inactivity logged successfully"})
    return JsonResponse({"status": "failure", "message": "Invalid request"}, status=400)


@user_passes_test(lambda user: is_in_group(user, "sales_team_leader") or is_in_group(user, "sales_manager"))
def sales_log_view(request):
    # Get the current user
    user = request.user

    # Check if the user is in the 'sales_manager' group
    if user.groups.filter(name='sales_manager').exists():
        # If the user is a sales manager, filter logs for users in 'sales_team_leader' or 'sales' groups
        sales_logs = SalesLog.objects.filter(
            Q(user__groups__name='sales_team_leader') | Q(user__groups__name='sales')
        ).order_by("-id")
        logs = sales_logs  # Combine the two querysets
    else:
        # Otherwise, filter logs based on the user's team
        team_members = UserLeader.objects.filter(leader=user).values_list('user', flat=True)
        logs = SalesLog.objects.filter(user__in=team_members).order_by("-id")  # Filter logs for team members



    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        logs = logs.filter(
            Q(message__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )

    # Date filter functionality
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    if start_date and end_date:
        logs = logs.filter(date__range=[start_date, end_date])

    # Pagination
    paginator = Paginator(logs, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'logs': page_obj,
        'search_query': search_query,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'sales_team_leader/sales_logs.html', context)


from django.utils import timezone


def send_cb_date_notifications():
    now = timezone.now()
    task_name = 'send_cb_date_notifications'

    try:
        task_log = TaskLog.objects.get(task_name=task_name)
        if task_log.last_run and task_log.last_run.date() == now.date():
            # Task has already run today, no need to run it again
            return
    except TaskLog.DoesNotExist:
        task_log = TaskLog(task_name=task_name)

    # Use timezone-aware dates for the current day's range
    today_start = now.astimezone().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = now.astimezone().replace(hour=23, minute=59, second=59, microsecond=999999)

    flags = TerminationCode.objects.filter(name__in=['CB', 'PR']).all()
    leads_with_cb_today = LeadTerminationCode.objects.filter(CB_date__range=(today_start, today_end), flag__in=flags)

    if not leads_with_cb_today:
        logger.info("No leads found with CB_date today.")
        return  # Exit the function if no leads are found

    for lead_termination in leads_with_cb_today:
        user = lead_termination.user
        
        if user is None:
            logger.warning("Lead termination has no associated user. Skipping.")
            continue
        
        # Attempt to send notification to the user
        try:
            Notification.objects.create(
                sender=user,
                receiver=user,
                message=f"You have a call back due today for lead '{lead_termination.lead.name}' in show '{lead_termination.sales_show.name}', Please Check your Prospect sheet",
                notification_type=5
            )
        except Exception as e:
            logger.error(f"Failed to create notification for user {user.username}: {e}")

        # Get sales managers and notify them
        sales_manager_group = Group.objects.get(name='sales_manager')
        sales_managers = sales_manager_group.user_set.all()
        
        for sales_manager in sales_managers:
            try:
                Notification.objects.create(
                    sender=user,
                    receiver=sales_manager,
                    message=f"{user.username} has a call back due today for lead '{lead_termination.lead.name}' in show '{lead_termination.sales_show.name}'",
                    notification_type=5
                )
            except Exception as e:
                logger.error(f"Failed to create notification for sales manager {sales_manager.username}: {e}")

        # Notify the user's leader
        try:
            user_leader = UserLeader.objects.get(user=user).leader
            Notification.objects.create(
                sender=user,
                receiver=user_leader,
                message=f"{user.username} has a call back due today for lead '{lead_termination.lead.name}' in show '{lead_termination.sales_show.name}'",
                notification_type=5
            )
        except UserLeader.DoesNotExist:
            logger.info(f"No leader found for user: {user.username}. Skipping notification to leader.")
        except Exception as e:
            logger.error(f"Failed to create notification for leader {user_leader.username}: {e}")

    # Update last_run after processing all notifications
    task_log.last_run = now
    task_log.save()



@login_required
def mark_as_read(request, notification_id):

    role = request.user.groups.first().name

    notification = get_object_or_404(Notification, id=notification_id, receiver=request.user)
    notification.read = True
    notification.save()
    return redirect(f'{role}:notifications')  # Redirect back to the notifications page



######################################################################################################################################################################################

## Function to import the sales inventory
# def import_lead_termination_history(request):
#     form = ImportSheetsForm()

#     if request.method == 'POST':
#         form = ImportSheetsForm(request.POST)
#         if form.is_valid():
#             folder_path = form.cleaned_data['folder_path']
#             if os.path.isdir(folder_path):
#                 files = os.listdir(folder_path)
#                 skipped_files = []
#                 errors = []

#                 def process_file(file):
#                     data = None
#                     file_extension = str(file).split('.')[-1]
#                     file_path = os.path.join(folder_path, file)

#                     try:
#                         if file_extension == 'xlsx':
#                             data = pd.read_excel(file_path, engine='openpyxl', header=None)
#                         elif file_extension == 'xls':
#                             data = pd.read_excel(file_path, header=None)
#                         elif file_extension == 'csv':
#                             data = pd.read_csv(file_path, header=None)
#                         else:
#                             return None  # Skip unsupported file types
#                     except Exception as e:
#                         logger.error(f"Error reading file {file}: {e}")
#                         return None  # Skip files that cannot be read

#                     # Ensure we have the required columns, except 'CB Date' which is optional
#                     required_columns = ['Termination Code', 'Special Notes']
#                     data.columns = data.iloc[0].tolist()  # Assuming first row contains column names
#                     data = data[1:]  # Skip header row

#                     # If the file doesn't have required columns, skip it
#                     if not all(col in data.columns for col in required_columns):
#                         logger.warning(f"File {file} is missing required columns: {', '.join(required_columns)}")
#                         skipped_files.append(file)
#                         return None

#                     # Create OldShow object for the file
#                     old_show, _ = OldShow.objects.get_or_create(name=file)

#                     for _, row in data.iterrows():
#                         try:
#                             # Fetch the lead from the database
#                             lead_name = row.get('Company Name', '').strip()
#                             lead = Lead.objects.filter(name__iexact=lead_name).first()
#                             if not lead:
#                                 logger.warning(f"Lead not found: {lead_name}")
#                                 continue

#                             # Fetch the TerminationCode object, skip if not found
#                             termination_code_str = row.get('Termination Code', '').strip()
#                             termination_code = TerminationCode.objects.filter(name=termination_code_str).first()
#                             if not termination_code:
#                                 logger.warning(f"Termination code not found: {termination_code_str} for lead {lead_name}. Skipping this lead.")
#                                 continue

#                             # CB Date is optional, so handle it separately
#                             cb_date = pd.to_datetime(row.get('CB Date', None), errors='coerce') if 'CB Date' in row else None

#                             # Special Notes
#                             notes = row.get('Special Notes', '').strip()

#                             # Create LeadTerminationHistory entry
#                             LeadTerminationHistory.objects.create(
#                                 lead=lead,
#                                 termination_code=termination_code,
#                                 cb_date=cb_date,
#                                 notes=notes,
#                                 old_show=old_show  # Save the OldShow reference
#                             )

#                         except Exception as e:
#                             logger.error(f"Error processing row: {row} - {e}")
#                             errors.append((file, str(e)))

#                     return old_show

#                 # Process files in parallel
#                 with ThreadPoolExecutor(max_workers=4) as executor:
#                     results = list(executor.map(process_file, files))

#                 # Log results and errors
#                 logger.info(f"Skipped {len(skipped_files)} files: {skipped_files}")
#                 logger.info(f"Encountered {len(errors)} errors: {errors}")

#                 return render(request, "leads/import_sheet.html", {
#                     "form": form,
#                     "success": "Sheets imported successfully."
#                 })
#             else:
#                 logger.error("The specified directory does not exist.")
#                 return render(request, "leads/import_sheet.html", {
#                     "form": form,
#                     "error": "The specified directory does not exist."
#                 })

#     return render(request, "leads/import_sheet.html", {
#         "form": form
#     })




## Function to upload the inventory to the database
# def import_sheets(request):
#     form = ImportSheetsForm()

#     if request.method == 'POST':
#         form = ImportSheetsForm(request.POST)
#         if form.is_valid():
#             folder_path = form.cleaned_data['folder_path']
#             if os.path.isdir(folder_path):
#                 files = os.listdir(folder_path)
#                 uploaded_sheets = []

#                 # Add lists to collect skipped files and errors
#                 skipped_files = []
#                 errors = []

#                 def process_file(file):
#                     data = None
#                     file_extension = str(file).split('.')[-1]
#                     file_path = os.path.join(folder_path, file)

#                     try:
#                         if file_extension == 'xlsx':
#                             data = pd.read_excel(file_path, engine='openpyxl', header=None)
#                         elif file_extension == 'xls':
#                             data = pd.read_excel(file_path, header=None)
#                         elif file_extension == 'csv':
#                             data = pd.read_csv(file_path, header=None)
#                         else:
#                             return None  # Skip unsupported file types
#                     except Exception as e:
#                         logger.error(f"Error reading file {file}: {e}")
#                         return None  # Skip files that cannot be read
                    
#                     try:
#                         if data.empty:
#                             logger.warning(f"Skipped empty file: {file}")
#                             skipped_files.append(file)
#                             return None  # Skip empty files
#                     except Exception as e:
#                         logger.error(f"Error processing file {file}: {e}")
#                         errors.append((file, str(e)))
#                         return None

#                     def ensure_str(value):
#                         return str(value) if pd.notna(value) else ''

#                     expected_headers = ['Company Name', 'Phone Number', 'Time Zone', 'Email', 'DM Name']
#                     extended_expected_headers = ['Company Name', 'Phone Number', 'Time Zone', 'Direct / Cell Number', 'Email', 'DM Name']
#                     first_row_as_str = data.iloc[0].apply(ensure_str)

#                     has_header = all(header.lower() in [col.lower() for col in first_row_as_str if col] for header in expected_headers)
#                     has_extended_header = all(header.lower() in [col.lower() for col in first_row_as_str if col] for header in extended_expected_headers)

#                     if has_header or has_extended_header:
#                         data.columns = data.iloc[0].apply(ensure_str).tolist()
#                         data = data[1:]

#                         if has_extended_header:
#                             for i, header in enumerate(extended_expected_headers):
#                                 if i >= len(data.columns) or data.columns[i] == '' or data.columns[i].lower() not in [h.lower() for h in extended_expected_headers]:
#                                     data.columns[i] = header
#                         else:
#                             for i, header in enumerate(expected_headers):
#                                 if i >= len(data.columns) or data.columns[i] == '' or data.columns[i].lower() not in [h.lower() for h in expected_headers]:
#                                     data.columns[i] = header
#                     else:
#                         num_cols = data.shape[1]
#                         default_columns = ['Company Name', 'Phone Number', 'Time Zone', 'Direct / Cell Number', 'Email', 'DM Name']
#                         data.columns = default_columns[:num_cols] + [f"Column_{i}" for i in range(num_cols - len(default_columns))]

#                     required_columns = ['Company Name', 'Phone Number', 'Email']
#                     missing_columns = [col for col in required_columns if col not in data.columns]
#                     if missing_columns:
#                         logger.warning(f"File {file} is missing required columns: {', '.join(missing_columns)}")
#                         return None

#                     data['Company Name'] = data['Company Name'].map(clean_company_name)
#                     data = data[data['Company Name'].apply(filter_companies)]

#                     # Ensure atomic transaction
#                     with transaction.atomic():
#                         sheet, created = Sheet.objects.get_or_create(
#                             name=file,
#                             defaults={'user': request.user, 'created_at': datetime.now()}
#                         )


#                     for _, row in data.iterrows():
#                         if not has_valid_contact(row):
#                             continue

#                         company_name = ensure_str(row.get('Company Name', ''))

#                         lead, created = Lead.objects.get_or_create(name=company_name)
#                         if created:
#                             lead.time_zone = ensure_str(row.get('Time Zone', ''))
#                             lead.save()

#                         phone_numbers = filter(None, [ensure_str(row.get('Phone Number', '')).strip(), ensure_str(row.get('Direct / Cell Number', '')).strip()])
#                         for phone_number in phone_numbers:
#                             if phone_number and is_valid_phone_number(phone_number):
#                                 # Avoid duplicates
#                                 LeadPhoneNumbers.objects.get_or_create(lead=lead, sheet=sheet, value=phone_number)

#                         email = ensure_str(row.get('Email', ''))
#                         if email:
#                             LeadEmails.objects.get_or_create(lead=lead, sheet=sheet, value=email)

#                         contact_name = ensure_str(row.get('DM Name', ''))
#                         if contact_name:
#                             LeadContactNames.objects.get_or_create(lead=lead, sheet=sheet, value=contact_name)

#                         # Ensure that the lead is linked to the current sheet
#                         if sheet not in lead.sheets.all():
#                             lead.sheets.add(sheet)

#                         sheet.is_approved = True
#                         sheet.save()
#                     return sheet

#                 with ThreadPoolExecutor(max_workers=4) as executor:
#                     results = list(executor.map(process_file, files))


#                 # After processing, check what was skipped or errored out
#                 logger.info(f"Skipped {len(skipped_files)} files: {skipped_files}")
#                 logger.info(f"Encountered {len(errors)} errors: {errors}")

#                 uploaded_sheets = [sheet for sheet in results if sheet]

#                 total_leads = sum(sheet.leads.count() for sheet in uploaded_sheets)
#                 Log.objects.create(message=f"[{request.user.username}] Uploaded [{total_leads}] total leads, [{len(uploaded_sheets)}] Sheets.")
#                 print(f"[{request.user.username}] Uploaded [{total_leads}] total leads, [{len(uploaded_sheets)}] Sheets.")

#                 return render(request, "leads/import_sheet.html", {
#                     "form": form,
#                     "success": "Sheets imported successfully."
#                 })
#             else:
#                 logger.error("The specified directory does not exist.")
#                 return render(request, "leads/import_sheet.html", {
#                     "form": form,
#                     "error": "The specified directory does not exist."
#                 })

#     return render(request, "leads/import_sheet.html", {
#         "form": form
#     })