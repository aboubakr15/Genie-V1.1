from datetime import datetime
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from main.custom_decorators import is_in_group
from .forms import AutoFillForm, UploadSheetsForm
from main.utils import get_lead_related_data, get_sheet_name, clean_company_name, filter_companies, send_websocket_message, NOTIFICATIONS_STATES
from main.models import Sheet, Lead, Notification, LatestSheet, LeadsAverage, LeadContactNames, LeadEmails, LeadPhoneNumbers, Log
import os, pandas as pd, logging, openpyxl
from django.contrib.auth.decorators import user_passes_test
from io import BytesIO
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl.styles import PatternFill
from django.contrib import messages
from IBH import settings
from django.core.files.storage import FileSystemStorage, default_storage
from django.db.models import OuterRef, Subquery


logger = logging.getLogger('custom')



@user_passes_test(lambda user: is_in_group(user, "leads"))
def index(request):
    query = request.GET.get('q')
    
    # Get the most recent phone number, email, and contact name
    recent_phone_number = LeadPhoneNumbers.objects.filter(lead=OuterRef('pk')).order_by('-id').values('value')[:1]
    recent_email = LeadEmails.objects.filter(lead=OuterRef('pk')).order_by('-id').values('value')[:1]
    recent_contact_name = LeadContactNames.objects.filter(lead=OuterRef('pk')).order_by('-id').values('value')[:1]

    if query:
        leads = Lead.objects.filter(name__icontains=query).order_by("-id")
    else:
        leads = Lead.objects.all().order_by("-id")
    
    # Annotate leads with recent contact information
    leads = leads.annotate(
        recent_phone_number=Subquery(recent_phone_number),
        recent_email=Subquery(recent_email),
        recent_contact_name=Subquery(recent_contact_name)
    )
    
    page = request.GET.get('page', '')
    paginator = Paginator(leads, 30)  

    try:
        leads_page = paginator.page(page)
    except PageNotAnInteger:
        leads_page = paginator.page(1)
    except EmptyPage:
        leads_page = paginator.page(paginator.num_pages)

    return render(request, 'leads/index.html', {
        'leads': leads_page,
        'query': query
    })
                 

@user_passes_test(lambda user: is_in_group(user, "leads"))
def auto_fill(request):
    
    form = AutoFillForm()

    if request.method == 'POST':
        form = AutoFillForm(request.POST, request.FILES)
        if 'file' in request.FILES and form.is_valid():
            file = request.FILES['file']
            latest_sheet = form.cleaned_data.get('latest_sheet')

            # Define the upload directory within the project folder
            upload_dir = os.path.join(settings.MEDIA_ROOT, 'auto_fill')

            # Create the directory if it does not exist
            os.makedirs(upload_dir, exist_ok=True)

            # Use FileSystemStorage to save the file
            fs = FileSystemStorage(location=upload_dir)
            filename = fs.save(file.name, file)
            
            file_extension = str(file).split('.')[-1]
            if file_extension == 'xlsx' or 'xls' or 'csv':
                # Create or get the sheet
                sheet, created = Sheet.objects.get_or_create(
                    name=filename,
                    defaults={'user': request.user, 'created_at': datetime.now()}
                )

                if latest_sheet:
                    LatestSheet.objects.get_or_create(
                        main_sheet=sheet,
                        latest_sheet=latest_sheet,
                        user=request.user
                    )   

            # Add a success message
            messages.success(request, "Sheet uploaded successfully, Please wait for approval.")

            # Create a notification with the uploaded sheets
            lead_user = request.user
            ops_tl = lead_user.leader.leader
            notification = Notification.objects.create(
                sender=lead_user,
                receiver=ops_tl,
                message=f'Auto fill request',
                notification_type=3, # 3: auto fill request
                read=False
            )
            notification.sheets.set([sheet])  # wrap sheet with [] to make it iter.
            notification.save()

            send_websocket_message(ops_tl.id, notification.id, notification.message, notification.read, NOTIFICATIONS_STATES['INFO'])

            return render(request, "leads/auto_fill.html", {
                "form": form,
            })
        else:
            messages.error(request, "Please upload a file.")
            return render(request, "leads/auto_fill.html", {
                "form": form,
            })
    else:
        return render(request, "leads/auto_fill.html", {
            "form": form,
        })


@user_passes_test(lambda user: is_in_group(user, "leads"))
def notifications(request):
    user = request.user
    notifications_for_user = Notification.objects.filter(
        receiver=user).order_by('-created_at')

    # Implement pagination
    page = request.GET.get('page', '')
    # Show 10 notifications per page
    paginator = Paginator(notifications_for_user, 5)

    try:
        notifications_page = paginator.page(page)
    except PageNotAnInteger:
        notifications_page = paginator.page(1)
    except EmptyPage:
        notifications_page = paginator.page(paginator.num_pages)

    return render(request, 'leads/notifications.html', {
        'notifications': notifications_page
    })


@user_passes_test(lambda user: is_in_group(user, "leads"))
def notification_detail(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id)
    sheets = notification.sheets.all()
    
    if not notification.read:
        notification.read = True
        notification.save()

    return render(request, 'leads/notification_detail.html',{
        'notification': notification,
        'sheets': sheets
    })


@user_passes_test(lambda user: is_in_group(user, "leads"))
def sheet_detail(request, notification_id, sheet_id):
    notification = get_object_or_404(Notification, id=notification_id)
    sheet = get_object_or_404(Sheet, id=sheet_id)
    leads = sheet.leads.all()
    return render(request, 'leads/sheet_detail.html', {'sheet': sheet, 'leads': leads, 'notification': notification})


@user_passes_test(lambda user: is_in_group(user, "leads"))
def upload_sheet(request):
    form = UploadSheetsForm()

    if request.method == 'POST':
        form = UploadSheetsForm(request.POST, request.FILES)
        if 'file' in request.FILES:
            file = request.FILES['file']

            # Define the upload directory within the project folder
            upload_dir = os.path.join(settings.MEDIA_ROOT, 'upload')

            # Create the directory if it does not exist
            os.makedirs(upload_dir, exist_ok=True)

            # Use FileSystemStorage to save the file
            fs = FileSystemStorage(location=upload_dir)
            filename = fs.save(file.name, file)
            
            file_extension = str(file).split('.')[-1]
            if file_extension == 'xlsx' or 'xls' or 'csv':
                # Create or get the sheet
                sheet, created = Sheet.objects.get_or_create(
                name=filename,
                defaults={'user': request.user, 'created_at': datetime.now()}
                )   

            # Add a success message
            messages.success(request, "Sheet imported successfully.")

            # Create a notification with the uploaded sheets
            lead_user = request.user
            ops_tl = lead_user.leader.leader
            notification = Notification.objects.create(
                sender=lead_user,
                receiver=ops_tl,
                message=f'sheet { sheet.name } uploaded',
                notification_type=0,
                read=False
            )
            notification.sheets.set([sheet])  # wrap sheet with [] to make it iter.
            notification.save()

            send_websocket_message(ops_tl.id, notification.id, notification.message,
                                   notification.read,  NOTIFICATIONS_STATES['INFO'])

            return render(request, "leads/upload_sheet.html", {
                "form": form,
            })
        else:
            messages.error(request, "Please upload a file.")
            return render(request, "leads/upload_sheet.html", {
                "form": form,
            })
    else:
        return render(request, "leads/upload_sheet.html", {
            "form": form,
        })


@user_passes_test(lambda user: is_in_group(user, "leads"))
def download_auto_fill_result(request, notification_id):
    # Get notification and sheet
    notification = Notification.objects.get(id=notification_id)
    sheet = notification.sheets.first()
    if not sheet:
        return HttpResponse("Sheet not found.", status=404)

    # Get the file path for the auto-fill result
    file_path = os.path.join(settings.MEDIA_ROOT, 'auto_fill', sheet.name)
    if not default_storage.exists(file_path):
        return HttpResponse("File not found.", status=404)

    # Load the file into a DataFrame
    with default_storage.open(file_path) as file:
        if file.name.endswith('.xlsx'):
            data = pd.read_excel(file, engine='openpyxl')
        elif file.name.endswith('.xls'):
            data = pd.read_excel(file)
        elif file.name.endswith('.csv'):
            data = pd.read_csv(file)
        else:
            return HttpResponse("Unsupported file format.", status=400)

    # Ensure the DataFrame contains the necessary columns
    required_columns = ['Company Name', 'Time Zone', 'Phone Number', 'Email', 'DM Name', 'Color']
    for col in required_columns:
        if col not in data.columns:
            data[col] = ''  # Initialize missing columns with empty values

    # Clean and filter company names
    data['Company Name'] = data['Company Name'].map(clean_company_name)
    data = data[data['Company Name'].apply(filter_companies)]
    company_names = data['Company Name'].unique()

    # Retrieve matching leads from the database
    leads = Lead.objects.filter(name__in=company_names)
    leads_dict = {lead.name: lead for lead in leads}

    # Calculate the number of autofilled leads
    num_leads_total = len(company_names)
    num_leads_autofilled = len(leads)

    # Check for the latest sheet
    latest_leads_dict = {}
    latest_sheet = None
    if sheet.name:
        try:
            latest_sheet_obj = LatestSheet.objects.get(main_sheet=sheet, user=request.user)
            latest_sheet = latest_sheet_obj.latest_sheet
        except LatestSheet.DoesNotExist:
            pass

        if latest_sheet:
            latest_leads = latest_sheet.leads.all()
            latest_leads_dict = {lead.name: lead for lead in latest_leads}

            # Add rows from the latest sheet to the data
            for lead in latest_leads:
                phone_number, email, contact_name = get_lead_related_data(lead)
                data = data[data['Company Name'] != lead.name]

                new_row = pd.DataFrame([{
                    'Company Name': lead.name,
                    'Time Zone': lead.time_zone,
                    'Phone Number': phone_number,
                    'Email': email,
                    'DM Name': contact_name,
                    'Color': 'green'
                }])

                new_row = new_row.reindex(columns=data.columns)
                data = pd.concat([data, new_row], ignore_index=True)

    # Fill the DataFrame with lead data from the database
    def fill_data(row):
        company_name = row.get('Company Name')
        if company_name not in latest_leads_dict:
            lead = leads_dict.get(company_name)
            if lead:
                phone_number, email, contact_name = get_lead_related_data(lead)
                row['Time Zone'] = lead.time_zone
                row['Phone Number'] = phone_number
                row['Email'] = email
                row['DM Name'] = contact_name
                row['Color'] = 'blue'
            else:
                row['Time Zone'] = ''
                row['Phone Number'] = ''
                row['Email'] = ''
                row['DM Name'] = ''
                row['Color'] = 'red'
        return row

    # Apply the fill_data function to the DataFrame
    data = data.apply(fill_data, axis=1)
    data.reset_index(drop=True, inplace=True)

    # Save the result in Excel format with color formatting
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        data.to_excel(writer, index=False, sheet_name='Leads')
        workbook = writer.book
        worksheet = writer.sheets['Leads']

        column_widths = {
            'Company Name': 20,
            'Time Zone': 15,
            'Phone Number': 20,
            'Email': 25,
            'DM Name': 20,
            'Color': 10
        }

        for col, width in column_widths.items():
            col_idx = data.columns.get_loc(col) + 1
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Define fills
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

        # Apply fills based on the color column
        for idx, row in data.iterrows():
            color = row['Color']
            fill = green_fill if color == 'green' else blue_fill if color == 'blue' else red_fill
            for col_num in range(1, len(row) + 1):
                worksheet.cell(row=idx + 2, column=col_num).fill = fill

    # Prepare the file for download
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="result_{sheet.name}.xlsx"'

    # Log the autofill result
    logger.info(f"[{request.user.username}] Made an AutoFill Request with {num_leads_total} total leads, {num_leads_autofilled} autofilled from the database.")
    Log.objects.create(message=f"[{request.user.username}] Made an AutoFill Request with {num_leads_total} total leads, {num_leads_autofilled} autofilled from the database.")

    # Clean up the file and delete the sheet record
    sheet.delete()
    os.remove(file_path)
    
    if latest_sheet:
        latest_sheet_obj.delete()

    return response
